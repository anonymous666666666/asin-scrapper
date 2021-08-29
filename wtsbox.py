import tkinter as tk
from tkinter import ttk
import pandas as pd
from googletrans import Translator
from selenium import webdriver
from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
from tkinter import messagebox
from tkinter import filedialog
from collections import OrderedDict
from tkinter import *


# google translate function
def googleTranslator(x):
    translator = Translator()
    out = translator.translate(x)
    return out.text


def askForFilePath():
    filetypes = (
        ('excel files', '*.xlsx'),
        ('All files', '*.*')
    )
    return filedialog.askopenfilename(filetypes=filetypes)


def BoxValue():
    return marketPlace.get()


def runMain():
    # Browse path

    filePath = askForFilePath()
    print(filePath)
    showImagePath(filePath)
    # load workbook
    print("Work in Progress...")
    wb = load_workbook(filePath)
    sheet1 = wb.worksheets[0]
    column = sheet1['A']  # Column
    row_count = sheet1.max_row
    column_list = [column[x].value for x in range(1, row_count)]

    # remove None values
    res = []
    for val in column_list:
        if val is not None:
            res.append(val)

    # remove whitespaces
    column_list_withoutSpace = [x.strip(' ') for x in res]

    # loop through all ASINs
    for s in range(len(column_list_withoutSpace)):
        if BoxValue() == "US":
            url = "https://www.amazon.com/dp/{0}{1}".format(column_list_withoutSpace[s], "?language=en_US")
            print(url)
        elif BoxValue() == "DE":
            url = "https://www.amazon.de/dp/{0}{1}".format(column_list_withoutSpace[s], "?language=en_GB")
            print(url)
        elif BoxValue() == "FR":
            url = "https://www.amazon.fr/dp/{}".format(column_list_withoutSpace[s])
            print(url)
        elif BoxValue() == "SG":
            url = "https://www.amazon.sg/dp/{}".format(column_list_withoutSpace[s])
            print(url)
        elif BoxValue() == "JP":
            url = "https://www.amazon.co.jp/dp/{0}{1}".format(column_list_withoutSpace[s], "?language=en_US")
            print(url)
        elif BoxValue() == "ES":
            url = "https://www.amazon.es/dp/{}".format(column_list_withoutSpace[s])
            print(url)
        elif BoxValue() == "UK":
            url = "https://www.amazon.co.uk/dp/{}".format(column_list_withoutSpace[s])
            print(url)
        elif BoxValue() == "CA":
            url = "https://www.amazon.ca/dp/{}".format(column_list_withoutSpace[s])
            print(url)
        else:
            url = "https://www.amazon.com/dp/{0}{1}".format(column_list_withoutSpace[s], "?language=en_CA")
            print(url)
        try:
            page = urlopen(url)
            soup = BeautifulSoup(page, 'html.parser')
            stockStatus = soup.find('div', {"id": "availability"})
            ratings = soup.find('div', {"id": "averageCustomerReviews"})
            headline = soup.find('div', {"id": "titleSection"})
            numRatings = soup.find('span', {"id": "acrCustomerReviewText"})
            price = soup.find('span', {"id": "priceblock_ourprice"})
            price2 = soup.find('div', {"id": "priceInsideBuyBox_feature_div"})
            article = stockStatus.text.strip()

            if not article:
                sheet1.cell(row=s + 2, column=2).value = "Lost Buy Box/ 3P seller only"
                try:
                    sheet1.cell(row=s + 2, column=3).value = headline.find('span').text.strip()
                except:
                    sheet1.cell(row=s + 2, column=3).value = "Headline not found"
                    print("Headline not found")
                try:
                    sheet1.cell(row=s + 2, column=4).value = price.text.strip()
                except:
                    sheet1.cell(row=s + 2, column=4).value = "Price not found"
                    print("Price not found")
                try:
                    sheet1.cell(row=s + 2, column=5).value = ratings.find('span').text.strip()
                except:
                    sheet1.cell(row=s + 2, column=5).value = "Ratings not found"
                    print("Ratings not found")
                try:
                    sheet1.cell(row=s + 2, column=6).value = numRatings.text.strip()
                except:
                    sheet1.cell(row=s + 2, column=6).value = "Number of reviews not found"
                    print("Number of reviews not found")

                print("{} - Lost Buy Box/ 3P seller only ".format(column_list_withoutSpace[s]))

            else:

                if BoxValue() == "FR" or BoxValue() == "ES":
                    sheet1.cell(row=s + 2, column=2).value = googleTranslator(article)
                else:
                    sheet1.cell(row=s + 2, column=2).value = article
                try:
                    sheet1.cell(row=s + 2, column=3).value = headline.find('span').text.strip()
                except:
                    sheet1.cell(row=s + 2, column=3).value = "Headline not found"
                    print("Headline not found")
                try:
                    sheet1.cell(row=s + 2, column=4).value = price.text.strip()
                except:
                    # try:
                    sheet1.cell(row=s + 2, column=4).value = price2.find('span').text.strip()
                    # except:
                    #     sheet1.cell(row=s + 2, column=5).value = "Price not found"
                    #     print("Price not found")
                try:
                    sheet1.cell(row=s + 2, column=5).value = ratings.find('span').text.strip()
                except:
                    sheet1.cell(row=s + 2, column=5).value = "Ratings not found"
                    print("Ratings not found")
                try:
                    sheet1.cell(row=s + 2, column=6).value = numRatings.text.strip()
                except:
                    sheet1.cell(row=s + 2, column=6).value = "Number of reviews not found"
                    print("Number of reviews not found")

            print("{} - Status added in sheet ".format(column_list_withoutSpace[s]))

        except:
            sheet1.cell(row=s + 2, column=2).value = "ASIN Page not found"
            print("Couldn't open this ASIN - {}".format(column_list_withoutSpace[s]))

    wb.save(filePath)
    messagebox.showinfo("showinfo", "execution complete")
    print("______________________________________")
    print("______________________________________")
    print("Work Done, Please check the excel file")


def grabAsins():
    driver = webdriver.Chrome()
    if not getInput():
        messagebox.showerror("Error", "Please Enter an Occasion page ID")
    else:
        driver.get("https://www.amazon.com/gcx/-/gfhz/events/?categoryId={}".format(getInput()))

        scrollToAsinWaterfall = driver.find_element_by_id("data-infinite-scroll")
        driver.execute_script("arguments[0].scrollIntoView();", scrollToAsinWaterfall)

        SCROLL_PAUSE_TIME = 6
        driver.execute_script("window.scrollBy(0, 1000)")
        time.sleep(SCROLL_PAUSE_TIME)
        last_height = driver.execute_script("return document.body.scrollHeight")
        asinList = []

        while True:
            ids = driver.find_element_by_id('data-infinite-scroll')
            var = ids.find_elements_by_tag_name('Section')
            y = len(var)
            # print(y)
            for i in range(0, y):
                try:
                    links = var[i].find_elements_by_tag_name('a')
                    x = len(links)

                    for a in range(0, x):
                        if links[a].get_attribute("href")[23:25] == "dp":
                            asinList.append(links[a].get_attribute("href")[26:36])

                except:
                    messagebox.showerror("Error", "Page couldn't load, Please try again")

            # Scroll down to bottom
            element = driver.find_element_by_id("navFooter")
            driver.execute_script("arguments[0].scrollIntoView();", element)
            time.sleep(SCROLL_PAUSE_TIME)

            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        finalAsins = list(OrderedDict.fromkeys(asinList))
        numberOfASINs = len(finalAsins)
        for asins in range(0, numberOfASINs):
            print(finalAsins[asins])

        try:
            AsinDataFrame = pd.DataFrame(finalAsins)
            writer = pd.ExcelWriter('asinList.xlsx', engine='xlsxwriter')
            AsinDataFrame.to_excel(writer, sheet_name='asinlist', index=False, header=False)
            writer.save()

            messagebox.showinfo("showinfo", "execution complete")
        except:

            messagebox.showerror("Error", "Please close asinList.xlsx file")


window = tk.Tk()
window.title('WTS')
window.geometry('400x350')
window.configure(bg='orange')

ttk.Label(window, text="ASIN DP scrapper",
          foreground="Black",
          font=("Times New Roman", 15), borderwidth=1, relief="solid").grid(row=0, column=1, padx=5, pady=15)

# label
ttk.Label(window, text="Select MP:",
          font=("Times New Roman", 10)).grid(column=0,
                                             row=5, padx=5, pady=15)

# Combobox creation
n = tk.StringVar()

marketPlace = ttk.Combobox(window, width=27, textvariable=n)

# Adding combobox drop down list
marketPlace['values'] = ('US',
                         'DE',
                         'FR',
                         'SG',
                         'JP',
                         'ES',
                         'UK',
                         'CA'
                         )

marketPlace.grid(column=1, row=5, padx=5, pady=15)
marketPlace.current()
n
button = tk.Button(window, text='I have already added ASINs in sheet', command=runMain, height=1, width=33)
button.grid(column=1, row=10, padx=5, pady=15)


def showImagePath(filePath):
    pathlabel = Label(window)
    pathlabel.config(text=filePath)
    pathlabel.grid(column=1, row=8, pady=10)


# button 2

def getInput():
    inp = inputtxt.get(1.0, "end-1c")
    # lbl.config(text="Provided Input: " + inp)
    if not inp:
        messagebox.showerror("Error", "Please Enter an Occasion page ID")
    else:
        return inp


# TextBox Creation
inputtxt = tk.Text(window,
                   height=1,
                   width=30)

# inputtxt.pack()
inputtxt.grid(column=1, row=7, padx=5, pady=15)
# Button Creation
printButton = tk.Button(window,
                        text="I want to scrape ASINs from the page",
                        command=grabAsins, height=1, width=33)

# Label Creation
ttk.Label(window, text="Enter Page ID:",
          font=("Times New Roman", 10)).grid(column=0, row=7,
                                             padx=5, pady=15)
# lbl.pack()
printButton.grid(column=1, row=9)

window.mainloop()
